import streamlit as st
import pandas as pd
import os
import re
import io
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set Page Config for premium layout (no emoji)
st.set_page_config(
    page_title="TME Music Royalty Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium CSS styling (glassmorphism details, font, colors)
st.markdown("""
<style>
    .reportview-container {
        background: #f0f2f6;
    }
    .metric-card {
        background-color: #ffffff;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border: 1px solid #eef2f6;
        text-align: center;
    }
    h1 {
        font-family: 'Outfit', 'Inter', sans-serif;
        color: #1e293b;
        font-weight: 700;
    }
    h2, h3 {
        font-family: 'Inter', sans-serif;
        color: #334155;
    }
</style>
""", unsafe_allow_html=True)

# Relative path definitions
ROYALTY_RATE_PATH = "../input/Royalty_Fee_Rate.xlsx"
SONG_REPORT_PATH = "../output/FORWARD/TME_Song_Report_Monthly_Forward.xlsx"

# Helper function to style Excel output for download
def style_excel_buffer(df):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Royalty_Report')
        
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    
    font_name = "Microsoft JhengHei"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Navy Blue
    zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") # Light ice-blue
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    max_row = ws.max_row
    max_col = ws.max_column
    
    col_types = {}
    for col_idx in range(1, max_col + 1):
        col_name = str(ws.cell(row=1, column=col_idx).value or '').lower()
        is_date_col = len(col_name) == 7 and col_name[4] == '-' and col_name[:4].isdigit() and col_name[5:].isdigit()
        if 'revenue' in col_name or 'royalty' in col_name or 'fee' in col_name or is_date_col:
            col_types[col_idx] = 'currency'
        elif 'share' in col_name or 'percentage' in col_name:
            col_types[col_idx] = 'percentage'
        elif 'isrc' in col_name:
            col_types[col_idx] = 'code'
        else:
            col_types[col_idx] = 'text'
            
    for r_idx in range(2, max_row + 1):
        ws.row_dimensions[r_idx].height = 20
        is_even = (r_idx % 2 == 0)
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
                
            c_type = col_types.get(c_idx, 'text')
            val = cell.value
            if isinstance(val, (int, float)):
                if c_type == 'currency':
                    cell.number_format = '#,##0.00'
                    cell.alignment = align_right
                elif c_type == 'percentage':
                    cell.number_format = '0.00%'
                    cell.alignment = align_right
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
            else:
                if c_type == 'code':
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format != 'General' and isinstance(cell.value, (int, float)):
                val_str += '    '
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 45)
        
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()

@st.cache_data
def load_base_data():
    if not os.path.exists(ROYALTY_RATE_PATH) or not os.path.exists(SONG_REPORT_PATH):
        return None, None
    df_rates = pd.read_excel(ROYALTY_RATE_PATH)
    df_report = pd.read_excel(SONG_REPORT_PATH)
    return df_rates, df_report

def main():
    st.title("TME Music Royalty Dashboard")
    st.markdown("---")
    
    df_rates, df_report = load_base_data()
    if df_rates is None or df_report is None:
        st.error("Royalty rate file or song report not found. Please ensure the pipeline and previous reports are completed.")
        return
        
    # Get available list of right holders
    df_rates['ISRC'] = df_rates['ISRC'].astype(str).str.strip().str.upper()
    df_report['ISRC'] = df_report['ISRC'].astype(str).str.strip().str.upper()
    right_holders = sorted(df_rates['right_holder'].dropna().unique().tolist())
    
    # Get available months
    months = sorted(list(set(re.match(r'^\d{4}-\d{2}', col).group(0) for col in df_report.columns if re.match(r'^\d{4}-\d{2}', col))))
    
    # Sidebar Filters
    st.sidebar.header("Filters")
    selected_holder = st.sidebar.selectbox("Select Right Holder:", right_holders)
    
    st.sidebar.markdown("---")
    start_month = st.sidebar.selectbox("Start Month:", months, index=0)
    end_month = st.sidebar.selectbox("End Month:", months, index=len(months)-1)
    
    if start_month > end_month:
        st.sidebar.error("Error: Start Month cannot be later than End Month!")
        return
        
    # Calculate Royalty data dynamically
    df_holder = df_rates[df_rates['right_holder'].str.strip() == selected_holder]
    df_holder_unique = df_holder.groupby('ISRC')['percentage'].first().reset_index()
    
    # Aggregate original song report by ISRC (handling duplicates)
    numeric_cols = [col for col in df_report.columns if col.endswith('_revenue') or col.endswith('_clicks') or col in ['total_revenue', 'total_click']]
    agg_dict = {col: 'sum' for col in numeric_cols}
    agg_dict['song'] = 'first'
    agg_dict['artist'] = 'first'
    df_report_grouped = df_report.groupby('ISRC').agg(agg_dict).reset_index()
    
    # Merge Rates & Song Report
    df_merged = pd.merge(df_holder_unique, df_report_grouped, on='ISRC', how='inner')
    
    if df_merged.empty:
        st.warning(f"No matched song data found for right holder '{selected_holder}'.")
        return
        
    # Calculate Royalty for selected months
    selected_months = [m for m in months if start_month <= m <= end_month]
    
    for m in selected_months:
        if f"{m}_revenue" in df_merged.columns:
            df_merged[f"{m}_royalty_fee"] = df_merged[f"{m}_revenue"] * df_merged['percentage']
            
    # Calculate total royalty fee in filtered timeframe
    df_merged['total_royalty_fee'] = df_merged[[f"{m}_royalty_fee" for m in selected_months]].sum(axis=1)
    
    # Filter columns to only show ISRC, Metadata, total royalty, and monthly royalty columns
    base_cols = ['ISRC', 'song', 'artist', 'total_royalty_fee']
    time_series_cols = [f"{m}_royalty_fee" for m in selected_months]
    
    final_cols = base_cols + time_series_cols
    final_cols = [col for col in final_cols if col in df_merged.columns]
    
    df_final = df_merged[final_cols].copy()
    
    # Rename monthly columns YYYY-MM_royalty_fee -> YYYY-MM
    rename_dict = {f"{m}_royalty_fee": m for m in selected_months}
    df_final = df_final.rename(columns=rename_dict)
    df_final = df_final.sort_values('total_royalty_fee', ascending=False).reset_index(drop=True)
    
    # ------------------ UI LAYOUT ------------------
    # Summary Metrics
    total_royalty = df_final['total_royalty_fee'].sum()
    unique_songs_count = len(df_final)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h4 style="color:#64748b;margin:0;">Right Holder</h4>
            <h2 style="color:#2563eb;margin:10px 0 0 0;font-weight:700;">{selected_holder}</h2>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h4 style="color:#64748b;margin:0;">Total Royalty Fee</h4>
            <h2 style="color:#16a34a;margin:10px 0 0 0;font-weight:700;">${total_royalty:,.2f}</h2>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h4 style="color:#64748b;margin:0;">Songs</h4>
            <h2 style="color:#ea580c;margin:10px 0 0 0;font-weight:700;">{unique_songs_count}</h2>
        </div>
        """, unsafe_allow_html=True)
        
    st.markdown("### Trend & Distribution Analysis")
    
    # Visualization Layout
    chart_col1, chart_col2 = st.columns([2, 1])
    
    with chart_col1:
        st.write("Monthly Royalty Fee Trend")
        monthly_trend = df_final[selected_months].sum().reset_index()
        monthly_trend.columns = ['YearMonth', 'Royalty Fee']
        st.line_chart(monthly_trend.set_index('YearMonth'), height=300)
        
    with chart_col2:
        st.write("Royalty Fee Contribution")
        pie_data = df_final[['song', 'total_royalty_fee']].copy()
        if len(pie_data) > 5:
            top_5 = pie_data.head(5)
            others_val = pie_data['total_royalty_fee'].iloc[5:].sum()
            others_row = pd.DataFrame([{'song': 'Others', 'total_royalty_fee': others_val}])
            pie_data = pd.concat([top_5, others_row], ignore_index=True)
        
        # Plot Pie Chart using Matplotlib
        fig, ax = plt.subplots(figsize=(6, 6))
        plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False
        
        ax.pie(
            pie_data['total_royalty_fee'],
            labels=pie_data['song'],
            autopct='%1.1f%%',
            startangle=140,
            colors=plt.cm.tab20c.colors
        )
        ax.axis('equal')
        fig.patch.set_alpha(0.0) # Transparent background
        st.pyplot(fig)

    # Royalty Details Table
    st.markdown("### Royalty Details")
    
    # Format display columns for st.dataframe
    format_dict = {'total_royalty_fee': '${:,.2f}'}
    for m in selected_months:
        format_dict[m] = '${:,.2f}'
        
    st.dataframe(
        df_final.style.format(format_dict),
        width="stretch",
        hide_index=True
    )
    
    # Export excel download button
    excel_data = style_excel_buffer(df_final)
    st.download_button(
        label="Download Excel File",
        data=excel_data,
        file_name=f"{selected_holder}_Royalty_Report_{start_month}_{end_month}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    main()
