"""
Reporting module for TME Music data aggregation and report export.
"""
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel_file(file_path: str):
    """Applies professional styling to an Excel workbook using openpyxl."""
    if not os.path.exists(file_path):
        return
        
    wb = load_workbook(file_path)
    
    # Styles config
    font_name = "Microsoft JhengHei"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Navy Blue
    zebra_fill = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid") # Very light ice-blue
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Keep gridlines visible
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
            
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Header formatting
        ws.row_dimensions[1].height = 26
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border
            
        # Data rows formatting
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Detect column types
        col_types = {}
        for col_idx in range(1, max_col + 1):
            col_name = str(ws.cell(row=1, column=col_idx).value or '').lower()
            if 'revenue' in col_name or '分成' in col_name or 'income' in col_name:
                col_types[col_idx] = 'currency'
            elif 'click' in col_name or '次數' in col_name or 'count' in col_name:
                col_types[col_idx] = 'clicks'
            elif 'share' in col_name or 'growth' in col_name or '成長' in col_name or '比例' in col_name:
                col_types[col_idx] = 'percentage'
            elif any(x in col_name for x in ['isrc', 'upc', 'date', 'year', 'month', '日期', '期間']):
                col_types[col_idx] = 'code'
            else:
                col_types[col_idx] = 'text'
                
        for r_idx in range(2, max_row + 1):
            ws.row_dimensions[r_idx].height = 20
            is_even = (r_idx % 2 == 0)
            
            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = thin_border
                
                if is_even:
                    cell.fill = zebra_fill
                    
                c_type = col_types.get(c_idx, 'text')
                val = cell.value
                
                if isinstance(val, (int, float)):
                    if c_type == 'currency':
                        cell.number_format = '#,##0.00'
                        cell.alignment = align_right
                    elif c_type == 'clicks':
                        cell.number_format = '#,##0'
                        cell.alignment = align_right
                    elif c_type == 'percentage':
                        if val > 1.0 or val < -1.0:
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = '0.00%'
                        cell.alignment = align_right
                    else:
                        cell.number_format = '#,##0'
                        cell.alignment = align_right
                else:
                    if c_type == 'code':
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left
                        
        # Auto column width
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format != 'General' and isinstance(cell.value, (int, float)):
                    val_str += '    '
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 45)
            
    wb.save(file_path)

# TODO: adjust column order
def generate_song_report(df_filtered: pd.DataFrame, output_path: str) -> pd.DataFrame:
    """Generate and save Song Revenue and Clicks report pivoted by YearMonth."""
    if df_filtered.empty:
        print("Warning: Input data is empty. Skipping song report.")
        return pd.DataFrame()

    # 1. Group by song attributes and month
    revenue_by_song_per_month = (
        df_filtered
        .groupby(['UPC', 'ISRC', 'YearMonth'])
        .agg(
            song=('standard_song', 'first'),
            artist=('standard_artist', 'first'),
            album=('Album', 'first'),
            revenue=('Revenue', 'sum'),
            clicks=('Total_Clicks', 'sum')
        )
        .reset_index()
    )
    
    # 2. Pivot table (MultiIndex columns: [clicks/revenue] x [YearMonth])
    song_pivot = revenue_by_song_per_month.pivot_table(
        index=['ISRC', 'song', 'artist', 'UPC', 'album'],
        columns='YearMonth',
        values=['clicks', 'revenue'],
        aggfunc='sum',
        fill_value=0
    )
    
    # 3. Calculate horizontal sums using MultiIndex slicing
    click_cols = [col for col in song_pivot.columns if col[0] == 'clicks']
    rev_cols = [col for col in song_pivot.columns if col[0] == 'revenue']
    
    song_pivot[('total_click', '')] = song_pivot[click_cols].sum(axis=1)
    song_pivot[('total_revenue', '')] = song_pivot[rev_cols].sum(axis=1)
    
    # Extract month list
    months = sorted(list(set(ym for metric, ym in song_pivot.columns if ym != '' and metric in ['clicks', 'revenue'])))
    
    # Reset index to move row index into columns
    song_pivot_df = song_pivot.reset_index()
    
    # 4. Flatten columns and reorder
    flat_cols = []
    for col in song_pivot_df.columns:
        level0, level1 = col
        if level1 == '':
            flat_cols.append(level0)
        else:
            flat_cols.append(f"{level1}_{level0}")
            
    song_pivot_df.columns = flat_cols
    
    # Build desired column order
    base_cols = ['ISRC', 'song', 'artist', 'UPC', 'album']
    summary_cols = ['total_click', 'total_revenue']
    time_cols = []
    for m in months:
        time_cols.extend([f"{m}_clicks", f"{m}_revenue"])
        
    final_cols = base_cols + summary_cols + time_cols
    song_final = song_pivot_df[final_cols].copy()
    
    # Sort by total revenue descending
    song_final = song_final.sort_values('total_revenue', ascending=False).reset_index(drop=True)
    
    # Export to Excel
    song_final.to_excel(output_path, index=False)
    style_excel_file(output_path)
    print(f"Song report exported successfully to: {output_path}")
    return song_final

def generate_album_report(df_filtered: pd.DataFrame, output_path: str) -> pd.DataFrame:
    """Generate and save Album Revenue report pivoted by YearMonth."""
    if df_filtered.empty:
        print("Warning: Input data is empty. Skipping album report.")
        return pd.DataFrame()
        
    # 1. Group by album attributes and month
    revenue_by_album_per_month = (
        df_filtered
        .groupby(['UPC', 'YearMonth'])
        .agg(
            album=('Album', 'first'),
            artist=('standard_artist', 'first'),
            revenue=('Revenue', 'sum')
        )
        .reset_index()
    )
    
    # 2. Pivot table (SingleIndex columns of YearMonth strings)
    album_pivot = revenue_by_album_per_month.pivot_table(
        index=['UPC', 'album', 'artist'],
        columns='YearMonth',
        values='revenue',
        aggfunc='sum',
        fill_value=0
    )
    
    # 3. Calculate row totals
    album_pivot['total_revenue'] = album_pivot.sum(axis=1)
    
    # Reset index to move row index into columns
    album_pivot_df = album_pivot.reset_index()
    
    # Sort months chronologically
    months = sorted([col for col in album_pivot.columns if col != 'total_revenue'])
    
    # Reorder columns: UPC, album, artist, total_revenue, then months
    final_cols = ['UPC', 'album', 'artist', 'total_revenue'] + months
    album_final = album_pivot_df[final_cols].copy()
    
    # Sort by total revenue descending
    album_final = album_final.sort_values('total_revenue', ascending=False).reset_index(drop=True)
    
    # Export to Excel
    album_final.to_excel(output_path, index=False)
    style_excel_file(output_path)
    print(f"Album report exported successfully to: {output_path}")
    return album_final

def generate_song_report_by_year(df_filtered: pd.DataFrame, output_path: str) -> pd.DataFrame:
    """Generate and save Song Revenue and Clicks report pivoted by Year."""
    if df_filtered.empty:
        print("Warning: Input data is empty. Skipping yearly song report.")
        return pd.DataFrame()

    # 1. Create a copy and extract Year (YYYY) from YearMonth (YYYY-MM)
    df = df_filtered.copy()
    if 'YearMonth' in df.columns:
        df['Year'] = df['YearMonth'].str.slice(0, 4)
    else:
        df['Year'] = 'Unknown'

    # 2. Group by song attributes (excluding UPC/Album as requested) and Year
    revenue_by_song_per_year = (
        df
        .groupby(['ISRC', 'Year'])
        .agg(
            song=('standard_song', 'first'),
            artist=('standard_artist', 'first'),
            revenue=('Revenue', 'sum'),
            clicks=('Total_Clicks', 'sum')
        )
        .reset_index()
    )
    
    # 3. Pivot table (MultiIndex columns: [clicks/revenue] x [Year])
    song_pivot = revenue_by_song_per_year.pivot_table(
        index=['ISRC', 'song', 'artist'],
        columns='Year',
        values=['clicks', 'revenue'],
        aggfunc='sum',
        fill_value=0
    )
    
    # 4. Calculate horizontal sums using MultiIndex slicing
    click_cols = [col for col in song_pivot.columns if col[0] == 'clicks']
    rev_cols = [col for col in song_pivot.columns if col[0] == 'revenue']
    
    song_pivot[('total_click', '')] = song_pivot[click_cols].sum(axis=1)
    song_pivot[('total_revenue', '')] = song_pivot[rev_cols].sum(axis=1)
    
    # Extract year list
    years = sorted(list(set(y for metric, y in song_pivot.columns if y != '' and metric in ['clicks', 'revenue'])))
    
    # Reset index to move row index into columns
    song_pivot_df = song_pivot.reset_index()
    
    # 5. Flatten columns and reorder
    flat_cols = []
    for col in song_pivot_df.columns:
        level0, level1 = col
        if level1 == '':
            flat_cols.append(level0)
        else:
            flat_cols.append(f"{level1}_{level0}")
            
    song_pivot_df.columns = flat_cols
    
    # Build desired column order (only including ISRC, song, artist, total metrics, and year metrics)
    base_cols = ['ISRC', 'song', 'artist']
    summary_cols = ['total_click', 'total_revenue']
    
    # Group all years' clicks first, then all years' revenue
    time_cols = [f"{y}_clicks" for y in years] + [f"{y}_revenue" for y in years]
        
    final_cols = base_cols + summary_cols + time_cols
    song_final = song_pivot_df[final_cols].copy()
    
    # Sort by total revenue descending
    song_final = song_final.sort_values('total_revenue', ascending=False).reset_index(drop=True)
    
    # Export to Excel
    song_final.to_excel(output_path, index=False)
    style_excel_file(output_path)
    print(f"Yearly song report exported successfully to: {output_path}")
    return song_final

