"""
Streamlit dashboard logic for the portal module.
Provides a unified company-internal dashboard for music asset management,
ingestion triggers, validation error exports, and formatted royalty reports.
"""
import os
import sys
import re
import io
import sqlite3
import pandas as pd
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Adjust sys.path to allow imports from parent directory
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from processor.etl_logic import run_etl_pipeline
from processor.error_reporter import get_error_count

# Page configuration
st.set_page_config(
    page_title="Royalty System Portal",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom css for a premium look
st.markdown("""
<style>
    .metric-card {
        background-color: #ffffff;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.05);
        border: 1px solid #eef2f6;
        text-align: center;
        margin-bottom: 15px;
    }
    .metric-card h4 {
        color: #64748b;
        margin: 0;
        font-size: 14px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .metric-card h2 {
        color: #1e293b;
        margin: 10px 0 0 0;
        font-weight: 700;
        font-size: 28px;
    }
    h1, h2, h3 {
        font-family: 'Inter', sans-serif;
        color: #1e293b;
    }
</style>
""", unsafe_allow_html=True)

LOCAL_DB_PATH = None

def get_db_connection():
    """Establish a connection to the SQLite database, using a local copy if in cloud storage."""
    global LOCAL_DB_PATH
    if LOCAL_DB_PATH is None:
        if config.DB_PATH.startswith("/data/"):
            LOCAL_DB_PATH = "/tmp/database.db"
            if os.path.exists(config.DB_PATH):
                import shutil
                try:
                    shutil.copy2(config.DB_PATH, LOCAL_DB_PATH)
                except Exception as e:
                    st.warning(f"Failed to copy GCS database to local scratch: {e}")
                    LOCAL_DB_PATH = config.DB_PATH
            else:
                pass
        else:
            LOCAL_DB_PATH = config.DB_PATH
            
    return sqlite3.connect(LOCAL_DB_PATH)

def style_excel_buffer(df: pd.DataFrame) -> bytes:
    """
    Format and style the DataFrame as a styled Excel sheet in bytes.
    Applies custom fonts, header fills, zebra-striping, borders, and number formatting.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Royalty_Statement')
        
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active
    
    font_name = "Microsoft JhengHei"
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_name, size=10)
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Light grey-blue
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    ws.views.sheetView[0].showGridLines = True
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    max_row = ws.max_row
    max_col = ws.max_column
    
    col_types = {}
    for col_idx in range(1, max_col + 1):
        col_name = str(ws.cell(row=1, column=col_idx).value or '').lower()
        is_date_col = len(col_name) == 7 and col_name[4] == '-' and col_name[:4].isdigit() and col_name[5:].isdigit()
        
        if 'revenue' in col_name or 'royalty' in col_name or 'fee' in col_name or is_date_col:
            col_types[col_idx] = 'currency'
        elif 'clicks' in col_name or 'click' in col_name or 'count' in col_name:
            col_types[col_idx] = 'clicks'
        elif 'percentage' in col_name or 'share' in col_name or 'split' in col_name:
            col_types[col_idx] = 'percentage'
        elif 'isrc' in col_name or 'upc' in col_name:
            col_types[col_idx] = 'code'
        else:
            col_types[col_idx] = 'text'
            
    for r_idx in range(2, max_row + 1):
        ws.row_dimensions[r_idx].height = 21
        is_even = (r_idx % 2 == 0)
        for c_idx in range(1, max_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border
            if is_even:
                cell.fill = zebra_fill
                
            c_type = col_types.get(c_idx, 'text')
            val = cell.value
            if isinstance(val, (int, float)):
                if c_type == 'currency':
                    cell.number_format = '$#,##0.00'
                    cell.alignment = align_right
                elif c_type == 'percentage':
                    cell.number_format = '0.00%'
                    cell.alignment = align_right
                elif c_type == 'clicks':
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                else:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
            else:
                if c_type == 'code':
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format != 'General' and isinstance(cell.value, (int, float)):
                val_str += '    '
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
        
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    return out_buf.getvalue()

def main():
    """Main dashboard entry point."""
    st.title("Royalty System Dashboard")
    st.markdown("---")
    
    conn = get_db_connection()
    
    # 1. Sidebar Panel - Data Filters First
    st.sidebar.subheader("Data Filters")
    
    # Fetch lists for dropdowns
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT right_holder FROM catalog_splits")
    right_holders = sorted([r[0] for r in cursor.fetchall() if r[0]])
    selected_holder = st.sidebar.selectbox("Select Right Holder:", ["All"] + right_holders)
    
    # Fetch list of unique songs based on selected right holder
    if selected_holder == "All":
        cursor.execute("SELECT DISTINCT song FROM etl_records ORDER BY song ASC")
    else:
        cursor.execute("""
            SELECT DISTINCT r.song 
            FROM etl_records r
            JOIN catalog_splits s ON r.isrc = s.isrc
            WHERE s.right_holder = ?
            ORDER BY r.song ASC
        """, (selected_holder,))
    song_list = sorted([r[0] for r in cursor.fetchall() if r[0]])
    selected_song = st.sidebar.selectbox("Select Song:", ["All Songs"] + song_list)
    
    # Months
    cursor.execute("SELECT DISTINCT year_month FROM etl_records ORDER BY year_month ASC")
    months = [r[0] for r in cursor.fetchall() if r[0]]
    
    if not months:
        st.warning("No records found in database. Please run the ETL Pipeline first.")
        conn.close()
        return
        
    default_start_idx = 0
    if "2024-01" in months:
        default_start_idx = months.index("2024-01")
        
    date_cols = st.sidebar.columns(2)
    with date_cols[0]:
        start_month = st.selectbox("Start Month:", months, index=default_start_idx)
    with date_cols[1]:
        end_month = st.selectbox("End Month:", months, index=len(months)-1)
    
    if start_month > end_month:
        st.sidebar.error("Error: Start Month cannot be later than End Month!")
        conn.close()
        return

    st.sidebar.markdown("---")
    st.sidebar.subheader("ETL Operations")
    
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0

    # File Uploader for End Users to upload spreadsheets directly to GCS
    uploaded_files = st.sidebar.file_uploader(
        "Upload raw Excel reports or ZIP", 
        type=["xlsx", "xls", "zip"], 
        accept_multiple_files=True,
        key=f"file_uploader_{st.session_state['uploader_key']}"
    )
    if uploaded_files:
        import zipfile
        success_count = 0
        errors = []
        for uploaded_file in uploaded_files:
            if uploaded_file.name.lower().endswith(".zip"):
                try:
                    zip_folder_name = os.path.splitext(uploaded_file.name)[0]
                    zip_target_dir = os.path.join(config.RAW_DATA_DIR, zip_folder_name)
                    os.makedirs(zip_target_dir, exist_ok=True)
                    with zipfile.ZipFile(io.BytesIO(uploaded_file.read())) as z:
                        for member in z.namelist():
                            filename = os.path.basename(member)
                            # Skip directories and macOS hidden files
                            if not filename or filename.startswith("._") or "__MACOSX" in member:
                                continue
                            if filename.lower().endswith((".xlsx", ".xls")):
                                target_path = os.path.join(zip_target_dir, filename)
                                try:
                                    with open(target_path, "wb") as f:
                                        f.write(z.read(member))
                                    success_count += 1
                                except Exception as inner_e:
                                    errors.append(f"Failed to extract {filename} from ZIP: {inner_e}")
                except Exception as e:
                    errors.append(f"Failed to open ZIP archive {uploaded_file.name}: {e}")
            else:
                target_path = os.path.join(config.RAW_DATA_DIR, uploaded_file.name)
                try:
                    with open(target_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    success_count += 1
                except Exception as e:
                    errors.append(f"Failed to save file {uploaded_file.name}: {e}")
                    
        if success_count > 0:
            st.sidebar.success(f"Successfully processed {success_count} file(s)!")
        if errors:
            for err in errors:
                st.sidebar.error(err)
                

    # Display ETL Success message if it was recently run
    if "etl_summary" in st.session_state:
        summary = st.session_state["etl_summary"]
        st.sidebar.success(
            f"ETL Done!\n"
            f"- Files: {summary['files_processed']}\n"
            f"- Records added: {summary['records_added']}\n"
            f"- Skipped: {summary['records_skipped_duplicate']}\n"
            f"- Errors logged: {summary['errors_logged']}"
        )
        del st.session_state["etl_summary"]

    # Trigger ETL Pipeline Button
    if st.sidebar.button("Analyze Data", use_container_width=True):
        with st.spinner("Executing ETL Pipeline..."):
            summary = run_etl_pipeline(conn, config.RAW_DATA_DIR)
            # Sync local copy back to GCS bucket to persist updates
            if LOCAL_DB_PATH == "/tmp/database.db" and config.DB_PATH.startswith("/data/"):
                import shutil
                try:
                    shutil.copy2(LOCAL_DB_PATH, config.DB_PATH)
                except Exception as e:
                    st.sidebar.error(f"Failed to save changes back to cloud storage: {e}")
            st.session_state["etl_summary"] = summary
            st.session_state["uploader_key"] += 1
            st.rerun()    

            
    # Export Error CSV Button
    err_count = get_error_count(conn)
    st.sidebar.markdown(f"Logged Errors: **{err_count}**")
    if err_count > 0:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT source_file, error_reason, COUNT(*) as error_count 
            FROM etl_errors 
            GROUP BY source_file, error_reason
            ORDER BY source_file ASC, error_count DESC
        """)
        df_errors = pd.DataFrame(cursor.fetchall(), columns=["Source File", "Reason", "Error Count"])
        
        st.sidebar.download_button(
            label="Download Error Report (CSV)",
            data=df_errors.to_csv(index=False).encode('utf-8-sig'),
            file_name="Error_Report.csv",
            mime="text/csv",
            use_container_width=True
        )

    # 2. Query Statistics & KPI Cards
    if selected_holder == "All":
        if selected_song == "All Songs":
            cursor.execute("SELECT SUM(revenue), SUM(clicks) FROM etl_records WHERE year_month >= ? AND year_month <= ?", (start_month, end_month))
        else:
            cursor.execute("SELECT SUM(revenue), SUM(clicks) FROM etl_records WHERE year_month >= ? AND year_month <= ? AND song = ?", (start_month, end_month, selected_song))
        row = cursor.fetchone()
        tot_rev = row[0] or 0.0
        tot_clicks = row[1] or 0.0
        
        if selected_song == "All Songs":
            cursor.execute("""
                SELECT SUM(r.revenue * s.percentage) 
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE r.year_month >= ? AND r.year_month <= ?
            """, (start_month, end_month))
        else:
            cursor.execute("""
                SELECT SUM(r.revenue * s.percentage) 
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE r.year_month >= ? AND r.year_month <= ? AND r.song = ?
            """, (start_month, end_month, selected_song))
        tot_royalty = cursor.fetchone()[0] or 0.0
        
        if selected_song == "All Songs":
            cursor.execute("SELECT COUNT(DISTINCT isrc) FROM etl_records WHERE year_month >= ? AND year_month <= ?", (start_month, end_month))
        else:
            cursor.execute("SELECT COUNT(DISTINCT isrc) FROM etl_records WHERE year_month >= ? AND year_month <= ? AND song = ?", (start_month, end_month, selected_song))
        tot_songs = cursor.fetchone()[0] or 0
    else:
        if selected_song == "All Songs":
            cursor.execute("""
                SELECT SUM(r.revenue), SUM(r.clicks), SUM(r.revenue * s.percentage)
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ?
            """, (selected_holder, start_month, end_month))
        else:
            cursor.execute("""
                SELECT SUM(r.revenue), SUM(r.clicks), SUM(r.revenue * s.percentage)
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ? AND r.song = ?
            """, (selected_holder, start_month, end_month, selected_song))
        row = cursor.fetchone()
        tot_rev = row[0] or 0.0
        tot_clicks = row[1] or 0.0
        tot_royalty = row[2] or 0.0
        
        if selected_song == "All Songs":
            cursor.execute("""
                SELECT COUNT(DISTINCT r.isrc)
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ?
            """, (selected_holder, start_month, end_month))
        else:
            cursor.execute("""
                SELECT COUNT(DISTINCT r.isrc)
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ? AND r.song = ?
            """, (selected_holder, start_month, end_month, selected_song))
        tot_songs = cursor.fetchone()[0] or 0
        
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Selected Holder</h4>
            <h2 style="color:#2563eb; font-size:18px; line-height:36px;">{selected_holder}</h2>
        </div>
        """, unsafe_allow_html=True)
    with col2:
        display_song = selected_song if len(selected_song) < 18 else selected_song[:15] + '...'
        st.markdown(f"""
        <div class="metric-card">
            <h4>Selected Song</h4>
            <h2 style="color:#2563eb; font-size:18px; line-height:36px;" title="{selected_song}">{display_song}</h2>
        </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Total Revenue</h4>
            <h2 style="color:#475569;">${tot_rev:,.2f}</h2>
        </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <h4>Unique Songs Count</h4>
            <h2 style="color:#ea580c;">{tot_songs}</h2>
        </div>
        """, unsafe_allow_html=True)

    # 3. Chart Visualizations
    st.markdown("### Trend Analysis")
    
    if selected_holder == "All":
        if selected_song == "All Songs":
            trend_query = """
                SELECT r.year_month, SUM(r.revenue * s.percentage) as royalty
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE r.year_month >= ? AND r.year_month <= ?
                GROUP BY r.year_month ORDER BY r.year_month ASC
            """
            trend_df = pd.read_sql_query(trend_query, conn, params=(start_month, end_month))
        else:
            trend_query = """
                SELECT r.year_month, SUM(r.revenue * s.percentage) as royalty
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE r.year_month >= ? AND r.year_month <= ? AND r.song = ?
                GROUP BY r.year_month ORDER BY r.year_month ASC
            """
            trend_df = pd.read_sql_query(trend_query, conn, params=(start_month, end_month, selected_song))
    else:
        if selected_song == "All Songs":
            trend_query = """
                SELECT r.year_month, SUM(r.revenue * s.percentage) as royalty
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ?
                GROUP BY r.year_month ORDER BY r.year_month ASC
            """
            trend_df = pd.read_sql_query(trend_query, conn, params=(selected_holder, start_month, end_month))
        else:
            trend_query = """
                SELECT r.year_month, SUM(r.revenue * s.percentage) as royalty
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ? AND r.song = ?
                GROUP BY r.year_month ORDER BY r.year_month ASC
            """
            trend_df = pd.read_sql_query(trend_query, conn, params=(selected_holder, start_month, end_month, selected_song))
        
    st.write("Monthly Revenue Trend")
    if not trend_df.empty:
        trend_df.columns = ['Year-Month', 'Allocated Royalty']
        # Reindex to fill in missing months with 0.0
        active_months = [m for m in months if start_month <= m <= end_month]
        trend_df = trend_df.set_index('Year-Month').reindex(active_months, fill_value=0.0)
        st.line_chart(trend_df, height=280)
    else:
        st.info("No timeline trend data to show.")
            
    # 4. Detailed Data Table & Excel Export
    st.markdown("### Detailed Statements")
    
    if selected_holder == "All":
        if selected_song == "All Songs":
            details_query = """
                SELECT 
                    isrc as ISRC,
                    song as Song,
                    artist as Artist,
                    album as Album,
                    upc as UPC,
                    year_month,
                    revenue,
                    clicks
                FROM etl_records
                WHERE year_month >= ? AND year_month <= ?
            """
            df_details = pd.read_sql_query(details_query, conn, params=(start_month, end_month))
        else:
            details_query = """
                SELECT 
                    isrc as ISRC,
                    song as Song,
                    artist as Artist,
                    album as Album,
                    upc as UPC,
                    year_month,
                    revenue,
                    clicks
                FROM etl_records
                WHERE year_month >= ? AND year_month <= ? AND song = ?
            """
            df_details = pd.read_sql_query(details_query, conn, params=(start_month, end_month, selected_song))
    else:
        if selected_song == "All Songs":
            details_query = """
                SELECT 
                    r.isrc as ISRC,
                    r.song as Song,
                    r.artist as Artist,
                    r.album as Album,
                    r.upc as UPC,
                    r.year_month,
                    (r.revenue * s.percentage) as royalty_fee
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ?
            """
            df_details = pd.read_sql_query(details_query, conn, params=(selected_holder, start_month, end_month))
        else:
            details_query = """
                SELECT 
                    r.isrc as ISRC,
                    r.song as Song,
                    r.artist as Artist,
                    r.album as Album,
                    r.upc as UPC,
                    r.year_month,
                    (r.revenue * s.percentage) as royalty_fee
                FROM etl_records r
                JOIN catalog_splits s ON r.isrc = s.isrc
                WHERE s.right_holder = ? AND r.year_month >= ? AND r.year_month <= ? AND r.song = ?
            """
            df_details = pd.read_sql_query(details_query, conn, params=(selected_holder, start_month, end_month, selected_song))
        
    if not df_details.empty:
        selected_months = sorted(df_details['year_month'].unique().tolist())
        
        if selected_holder == "All":
            df_pivot = df_details.pivot_table(
                index=['ISRC', 'Song', 'Artist', 'Album', 'UPC'],
                columns='year_month',
                values=['clicks', 'revenue'],
                aggfunc='sum',
                fill_value=0.0
            )
            
            # Calculate horizontal sums
            click_cols = [col for col in df_pivot.columns if col[0] == 'clicks']
            rev_cols = [col for col in df_pivot.columns if col[0] == 'revenue']
            
            df_pivot[('clicks', 'Total Clicks')] = df_pivot[click_cols].sum(axis=1)
            df_pivot[('revenue', 'Total Revenue')] = df_pivot[rev_cols].sum(axis=1)
            
            df_pivot_df = df_pivot.reset_index()
            
            # Flatten MultiIndex columns
            flat_cols = []
            for col in df_pivot_df.columns:
                metric, level1 = col
                if level1 == '':
                    flat_cols.append(metric)
                elif level1 == 'Total Clicks':
                    flat_cols.append('Total Clicks')
                elif level1 == 'Total Revenue':
                    flat_cols.append('Total Revenue')
                else:
                    flat_cols.append(f"{level1} {metric.capitalize()}")
            df_pivot_df.columns = flat_cols
            
            base_cols = ['ISRC', 'Song', 'Artist', 'Album', 'UPC']
            summary_cols = ['Total Clicks', 'Total Revenue']
            time_cols = []
            for m in selected_months:
                time_cols.extend([f"{m} Clicks", f"{m} Revenue"])
                
            final_cols = base_cols + summary_cols + time_cols
            df_final = df_pivot_df[final_cols].sort_values('Total Revenue', ascending=False).reset_index(drop=True)
            
            format_dict = {'Total Revenue': '${:,.2f}', 'Total Clicks': '{:,.0f}'}
            for m in selected_months:
                format_dict[f"{m} Revenue"] = '${:,.2f}'
                format_dict[f"{m} Clicks"] = '{:,.0f}'
        else:
            df_pivot = df_details.pivot_table(
                index=['ISRC', 'Song', 'Artist', 'Album', 'UPC'],
                columns='year_month',
                values='royalty_fee',
                aggfunc='sum',
                fill_value=0.0
            ).reset_index()
            
            df_pivot['Total Royalty'] = df_pivot[selected_months].sum(axis=1)
            
            final_cols = ['ISRC', 'Song', 'Artist', 'Album', 'UPC', 'Total Royalty'] + selected_months
            df_final = df_pivot[final_cols].sort_values('Total Royalty', ascending=False).reset_index(drop=True)
            
            format_dict = {'Total Royalty': '${:,.2f}'}
            for m in selected_months:
                format_dict[m] = '${:,.2f}'
                
        st.dataframe(
            df_final.style.format(format_dict),
            use_container_width=True,
            hide_index=True
        )
        
        excel_data = style_excel_buffer(df_final)
        st.download_button(
            label="Download Excel Report",
            data=excel_data,
            file_name=f"Report_{selected_holder.replace(' ', '_')}_{start_month}_{end_month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_btn"
        )
    else:
        st.info("No matching records found for the selected timeline filters.")
        
    conn.close()

if __name__ == "__main__":
    main()
